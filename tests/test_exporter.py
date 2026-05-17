from __future__ import annotations

import json
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from ev_charge_testcase_generator.exporter import (
    EXECUTABLE_EXCEL_HEADERS,
    export_cases_to_excel,
    export_cases_to_json,
    export_to_excel,
    export_to_json,
)
from ev_charge_testcase_generator.models import (
    Assertion,
    CleanupStep,
    ExecutableStep,
    ExecutableTestCase,
    Precondition,
)


TMP_EXPORT_DIR = Path(__file__).resolve().parent / "_tmp_exporter"


@pytest.fixture(autouse=True)
def clean_export_tmp_dir():
    if TMP_EXPORT_DIR.exists():
        shutil.rmtree(TMP_EXPORT_DIR)
    yield
    if TMP_EXPORT_DIR.exists():
        shutil.rmtree(TMP_EXPORT_DIR)


def _output_path(name: str) -> Path:
    TMP_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = TMP_EXPORT_DIR / name
    if path.exists():
        path.unlink()
    return path


def _build_case(case_id: str = "TC-DC-NORMAL-001") -> ExecutableTestCase:
    return ExecutableTestCase(
        case_id=case_id,
        case_name="直流充电握手阶段BHM响应测试",
        scene_type="DC",
        condition_type="normal",
        test_type="positive",
        standard_source="GB/T 34658-2025",
        test_stage="低压辅助上电及充电握手阶段",
        target_object="BMS/EVCC",
        preconditions=[Precondition(condition_id="PRE-001", description="完成物理连接")],
        steps=[
            ExecutableStep(
                step_id=1,
                action_id="SEND_CHM-00005",
                action_name="发送CHM报文",
                action_type="发送报文",
                message="CHM",
                parameters={"period_ms": "250"},
            )
        ],
        assertions=[
            Assertion(
                assertion_id="CHECK_BHM_CONTENT-00008",
                assertion_type="message",
                description="检查BHM报文内容",
                message="BHM",
            )
        ],
        cleanup_steps=[CleanupStep(step_id=1, action_id="CAN_RECORD_STOP-00023", action_name="停止CAN记录")],
        raw_requirement="测试直流充电过程中BMS是否回复BHM报文",
    )


def test_export_to_json() -> None:
    output_path = _output_path("case.json")

    export_to_json(_build_case(), output_path)

    data = json.loads(output_path.read_text(encoding="utf-8"))
    assert data["case_id"] == "TC-DC-NORMAL-001"
    assert data["steps"][0]["parameters"] == {"period_ms": "250"}


def test_export_to_excel() -> None:
    output_path = _output_path("case.xlsx")

    export_to_excel(_build_case(), output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    assert headers[-len(EXECUTABLE_EXCEL_HEADERS) :] == EXECUTABLE_EXCEL_HEADERS
    assert worksheet["I2"].value == "precondition"
    assert worksheet["I3"].value == "step"


def test_export_cases_to_json() -> None:
    test_cases = [_build_case("TC-DC-NORMAL-001"), _build_case("TC-DC-NORMAL-002")]
    output_dir = TMP_EXPORT_DIR / "json_batch"
    output_dir.mkdir(parents=True, exist_ok=True)

    output_paths = export_cases_to_json(test_cases, output_dir)

    assert len(output_paths) == 2
    assert {path.name for path in output_paths} == {"TC-DC-NORMAL-001.json", "TC-DC-NORMAL-002.json"}


def test_export_cases_to_excel() -> None:
    output_path = _output_path("cases.xlsx")
    test_cases = [_build_case("TC-DC-NORMAL-001"), _build_case("TC-DC-NORMAL-002")]

    export_cases_to_excel(test_cases, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook.active
    assert worksheet.max_row == 9
    assert worksheet["A2"].value == "TC-DC-NORMAL-001"
    assert worksheet["A6"].value == "TC-DC-NORMAL-002"
